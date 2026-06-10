import base64, io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "LP Tracker Jun 2026"

def fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
def fnt(h, sz=11, bold=False, italic=False):
    return Font(color=h.lstrip("#"), size=sz, bold=bold, italic=italic, name="Arial")
def bb(h="E8ECF3"):
    s = Side(style="thin", color=h.lstrip("#"))
    return Border(bottom=s)
def mid(): return Alignment(vertical="center")
def mc():  return Alignment(vertical="center", horizontal="center")
def mr():  return Alignment(vertical="center", horizontal="right")

CAT = {
    "QFLAVOURS":   ("E1F5EE","04342C","1D9E75"),
    "ACCOUNTING":  ("E6F1FB","042C53","378ADD"),
    "ZATCA":       ("EEEDFE","26215C","7F77DD"),
    "POS":         ("FAECE7","4A1B0C","D85A30"),
    "SECTOR":      ("FAEEDA","412402","BA7517"),
    "BOOKKEEPING": ("EAF3DE","173404","639922"),
    "OTHER":       ("F1EFE8","2C2C2A","888780"),
    "SUMMARY":     ("F5F7FB","444441","888780"),
}
STATUS = {"Live":("EAF3DE","27500A"), "Draft":("F1EFE8","5F5E5A")}
TMPL   = {"qnav":("E6F1FB","0C447C"),"qnav-dark":("EEEDFE","3C3489"),
          "qnav-mega":("E1F5EE","085041"),"old-nav":("F1EFE8","5F5E5A")}
FIX    = {"N/A":("F1EFE8","5F5E5A"),"Needed":("FAEEDA","633806"),"Done":("EAF3DE","27500A")}

ROWS = [
    ("HDR","#","Category","Page (EN)","Live URL","Preview URL","Status","Template","Mobile Fix","Total Subs","May 2026","Jun 2026","WP Modified","Notes"),
    ("CAT","QFLAVOURS"),
    ("DAT","1","Qflavours","Qflavours — Cashier & Restaurant","https://lp.qoyod.com/qflavours/","https://lp.qoyod.com/?page_id=8&preview=true","Live","old-nav","N/A","0","—","—","2026-05-25","Form 6dbef1e"),
    ("DAT","2","Qflavours","Qflavours — Know page","https://lp.qoyod.com/know-qflavours/","https://lp.qoyod.com/?page_id=240&preview=true","Live","old-nav","N/A","0","—","—","2026-05-25","Revamped 2026-05-25"),
    ("DAT","3","Qflavours","Flavours POS for Restaurants","","https://lp.qoyod.com/?page_id=558&preview=true","Draft","—","—","—","—","—","2026-05-23",""),
    ("DAT","4","Qflavours","Flavours POS Rebuilt (v2)","","https://lp.qoyod.com/?page_id=840&preview=true","Draft","—","—","—","—","—","2026-05-25","Rebuilt 2026-05-25"),
    ("CAT","ACCOUNTING"),
    ("DAT","5","Accounting","Accounting — best solution","https://lp.qoyod.com/accounting/","https://lp.qoyod.com/?page_id=303&preview=true","Live","old-nav","N/A","77","67","10","2026-06-08","Legacy live — highest traffic"),
    ("DAT","6","Accounting","Accounting for Accountants","","https://lp.qoyod.com/?page_id=553&preview=true","Draft","qnav","Needed","—","—","—","2026-05-25","Rebuilt 2026-05-26"),
    ("DAT","7","Accounting","Cloud Accounting Software","https://lp.qoyod.com/accounting-system/","https://lp.qoyod.com/?page_id=850&preview=true","Live","qnav","Done 2026-06-09","4","4","0","2026-06-09","A/B base page"),
    ("CAT","ZATCA"),
    ("DAT","8","ZATCA","E-Invoice Integration","https://lp.qoyod.com/einvoice-integration/","https://lp.qoyod.com/?page_id=463&preview=true","Live","old-nav","N/A","11","4","7","2026-06-08","Legacy live"),
    ("DAT","9","ZATCA","ZATCA E-Invoice Phase 2","https://lp.qoyod.com/zatca-einvoice/","https://lp.qoyod.com/?page_id=851&preview=true","Live","qnav-dark","Done 2026-06-09","2","0","2","2026-06-09","Do NOT batch-upload"),
    ("CAT","POS"),
    ("DAT","10","POS / Retail","POS for Retail with ZATCA","","https://lp.qoyod.com/?page_id=548&preview=true","Draft","—","—","—","—","—","2026-05-23",""),
    ("DAT","11","POS / Retail","Accounting for Retail","","https://lp.qoyod.com/?page_id=559&preview=true","Draft","qnav","Needed","—","—","—","2026-05-25","Rebuilt 2026-05-26"),
    ("DAT","12","POS / Retail","Qoyod POS for Retail","","https://lp.qoyod.com/?page_id=841&preview=true","Draft","—","—","—","—","—","2026-05-23",""),
    ("CAT","SECTOR"),
    ("DAT","13","Sector","Accounting for Services","","https://lp.qoyod.com/?page_id=682&preview=true","Draft","qnav","Needed","—","—","—","2026-05-25","Rebuilt 2026-05-26"),
    ("DAT","14","Sector","Accounting for Tech Companies","https://lp.qoyod.com/tech-sector/","https://lp.qoyod.com/?page_id=683&preview=true","Live","qnav","Done 2026-06-09","2","0","2","2026-06-09",""),
    ("DAT","15","Sector","Accounting for Real Estate","","https://lp.qoyod.com/?page_id=685&preview=true","Draft","—","—","—","—","—","2026-05-23",""),
    ("CAT","BOOKKEEPING"),
    ("DAT","16","Bookkeeping","Bookkeeping for Business Owners","","https://lp.qoyod.com/?page_id=557&preview=true","Draft","qnav","Needed","—","—","—","2026-06-07","Rebuilt 2026-05-26"),
    ("CAT","OTHER"),
    ("DAT","17","Other","Qawaem — File Financial Statements","https://lp.qoyod.com/qawaem/","https://lp.qoyod.com/?page_id=773&preview=true","Live","qnav-mega","Done 2026-06-08","4","2","2","2026-06-08","All 6 mobile fixes"),
    ("CAT","SUMMARY"),
    ("SUM","Live pages","8","303 / 463 / 683 / 773 / 850 / 851 / 8 / 240"),
    ("SUM","Draft pages","9","548 / 553 / 557 / 558 / 559 / 682 / 685 / 840 / 841"),
    ("SUM","Mobile Fix Done","4","tech-sector / accounting-system / zatca-einvoice / qawaem"),
    ("SUM","Mobile Fix Needed","4","accounting-accountants / retail-accounting / services-sector / bookkeeping-smb"),
]

COL_W = [5, 14, 32, 28, 28, 9, 12, 18, 12, 9, 9, 14, 34]
cc = None

for row_data in ROWS:
    rtype = row_data[0]
    cells = list(row_data[1:])

    if rtype == "HDR":
        ws.append(cells)
        r = ws.max_row
        ws.row_dimensions[r].height = 22
        for ci, cell in enumerate(ws[r], 1):
            cell.fill = fill("021544")
            cell.font = fnt("FFFFFF", sz=10, bold=True)
            cell.alignment = mc() if ci in (1,6,7,8,9,10,11) else mid()

    elif rtype == "CAT":
        cc = CAT.get(cells[0], CAT["OTHER"])
        ws.append([f"  {cells[0].title()}"])
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"A{r}:M{r}")
        for ci in range(1, 14):
            ws.cell(r, ci).fill = fill(cc[0])
        ws.cell(r, 1).font = fnt(cc[1], sz=9, bold=True)
        ws.cell(r, 1).alignment = mid()

    elif rtype == "DAT":
        ws.append(cells)
        r = ws.max_row
        ws.row_dimensions[r].height = 20
        for ci in range(1, 14):
            cell = ws.cell(r, ci)
            cell.fill = fill("FFFFFF")
            cell.border = bb()
            cell.alignment = mid()
        ws.cell(r,1).font = fnt("B4B2A9", sz=10)
        ws.cell(r,1).alignment = mc()
        if cc: ws.cell(r,2).font = fnt(cc[2], sz=10, bold=True)
        ws.cell(r,3).font = fnt("0B1220", sz=11, bold=True)
        for ci, url in [(4, cells[3]), (5, cells[4])]:
            cell = ws.cell(r, ci)
            if url.startswith("http"):
                cell.hyperlink = url
                cell.font = Font(color="185FA5", size=10, underline="single", name="Arial")
            else:
                cell.font = fnt("B4B2A9", sz=10)
        st = cells[5]
        if st in STATUS:
            ws.cell(r,6).fill = fill(STATUS[st][0])
            ws.cell(r,6).font = fnt(STATUS[st][1], sz=10, bold=True)
            ws.cell(r,6).alignment = mc()
        tmpl = cells[6]
        if tmpl in TMPL:
            ws.cell(r,7).fill = fill(TMPL[tmpl][0])
            ws.cell(r,7).font = fnt(TMPL[tmpl][1], sz=10, bold=True)
            ws.cell(r,7).alignment = mc()
        fk = cells[7].split(" ")[0].strip()
        if fk in FIX:
            ws.cell(r,8).fill = fill(FIX[fk][0])
            ws.cell(r,8).font = fnt(FIX[fk][1], sz=10, bold=True)
            ws.cell(r,8).alignment = mc()
        for ci in [9,10,11]:
            v = cells[ci-1]
            cell = ws.cell(r, ci)
            cell.alignment = mr()
            if v and v != "—" and v != "0":
                cell.font = fnt("021544", sz=11, bold=True)
            else:
                cell.font = fnt("B4B2A9", sz=10)
        ws.cell(r,12).font = fnt("888780", sz=10)
        ws.cell(r,13).font = fnt("5F5E5A", sz=10, italic=True)

    elif rtype == "SUM":
        ws.append(cells)
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        for ci in range(1, 14):
            cell = ws.cell(r, ci)
            cell.fill = fill("F5F7FB")
            cell.font = fnt("444441", sz=10, bold=True)
            cell.alignment = mid()

ws.freeze_panes = "A2"
for i, w in enumerate(COL_W, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

buf = io.BytesIO()
wb.save(buf)
with open("D:/Nexa Performance Agent/lp_tracker_formatted.xlsx", "wb") as f:
    f.write(buf.getvalue())
b64 = base64.b64encode(buf.getvalue()).decode()
print("OK len=" + str(len(b64)))
print(b64)
